import openpyxl

def save_blockchain(blockchain, file_name):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Write block data to excel
    for i, block in enumerate(blockchain.chain):
        sheet.cell(row=i+1, column=1, value=block.index)
        sheet.cell(row=i+1, column=2, value=block.timestamp)
        sheet.cell(row=i+1, column=3, value=block.data)
        sheet.cell(row=i+1, column=4, value=block.previous_hash)
        sheet.cell(row=i+1, column=5, value=block.hash)
        sheet.cell(row=i+1, column=6, value=block.nonce)

    wb.save(file_name)

def open_blockchain(file_name, blockchain, Block):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        block = Block(
            index=sheet.cell(row=i, column=1).value,
            timestamp=sheet.cell(row=i, column=2).value,
            data=sheet.cell(row=i, column=3).value,
            previous_hash=sheet.cell(row=i, column=4).value
        )
        block.hash = sheet.cell(row=i, column=5).value
        block.nonce = sheet.cell(row=i, column=6).value
        blockchain.chain.append(block)
    return blockchain